"""The daily report workbook's chart sheets: `Report` and the five community
sheets, plus the pieces every chart shares.

Layer: services. Split out of `work_order_report_xlsx.py` so neither module
crosses the 500-line rule (redesign spec E12, plan Task 6). The dependency runs
one way -- the renderer imports from here, never the reverse -- so the split
adds no import cycle. Everything here is still a pure function of a
`DailyReport` (E13): no queries, no clock.

Spec: docs/superpowers/specs/2026-08-30-hub-report-xlsx-redesign-design.md

**Every chart reads the hidden `Chart Data` sheet (E7)**, one labelled block
per chart written by `_ChartData`'s cursor, with `visible_cells_only = False`
so Excel plots it. The designed sheets carry the same numbers as styled
tables, from the same payload -- never as chart sources.

**No pie of zeros (§4.2).** `_pie` writes the empty-state line instead when a
group has no rows; a pie with no area is a rendering bug, not a data point.
"""

from __future__ import annotations

from decimal import Decimal
from typing import Optional

from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.domain import labor_day
from app.services import _xlsx_theme as theme
from app.services.work_order_report import (
    CLOSING_STATUSES,
    STATUS_LABELS,
    DailyReport,
    ReportRow,
)
from app.services.work_order_report_buckets import (
    BUCKET_CLOSED,
    BUCKET_KEYS,
    BUCKET_LABELS,
    CommunityDistribution,
    grid_of,
    primary_community,
)

PLACEHOLDER = "(none)"
EMPTY_STATE = "No live or recently closed work orders."
EMPTY_COMMUNITY_STATE = "No live or recently closed work orders in this community."

BUCKET_COLOR_ORDER = [theme.BUCKET_COLORS[key] for key in BUCKET_KEYS]


# --------------------------------------------------------------------------
# Shared pieces
# --------------------------------------------------------------------------


class _ChartData:
    """Cursor over the hidden source sheet: one named block per chart (E7).

    Each block is a one-cell name, a header row, then its rows, then a blank
    row -- so a person who unhides the sheet can find what a chart reads."""

    def __init__(self, sheet: Worksheet) -> None:
        self.sheet = sheet
        self.row = 1

    def block(self, name: str, header: list, rows: list[list]) -> tuple[int, int]:
        """Write the block; return `(header_row, last_row)` for a `Reference`."""
        self.sheet.cell(row=self.row, column=1, value=name)
        header_row = self.row + 1
        for offset, value in enumerate(header, start=1):
            self.sheet.cell(row=header_row, column=offset, value=value)
        for index, values in enumerate(rows, start=1):
            for offset, value in enumerate(values, start=1):
                self.sheet.cell(row=header_row + index, column=offset, value=value)
        last_row = header_row + len(rows)
        self.row = last_row + 2
        return header_row, last_row


def _population_caveat(payload: DailyReport) -> str:
    """§2.1, on every designed sheet: what the population is, and why the
    community totals do not sum to the company total."""
    generated = payload.generated_at.astimezone(labor_day.CENTRAL)
    week = payload.week
    return (
        f"Live work orders as of {generated:%Y-%m-%d %H:%M} Central, plus work "
        f"orders closed {week.start.isoformat()} – {week.end.isoformat()}. A work "
        "order named in two communities is counted in both; community totals do "
        "not sum to the company total."
    )


def _bucket_rows(group: CommunityDistribution) -> list[list]:
    """Label / count / share, in bucket order. Shares are real fractions."""
    return [
        [
            BUCKET_LABELS[key],
            group.counts[key],
            (group.counts[key] / group.total) if group.total else 0,
        ]
        for key in BUCKET_KEYS
    ]


def _pie(
    sheet: Worksheet,
    chart_data: _ChartData,
    *,
    name: str,
    title: str,
    group,
    cells: str,
    legend: Optional[str],
    percent_labels: bool,
    empty_text: str = EMPTY_STATE,
) -> None:
    """A bucket pie for `group` (anything with `.total` and `.counts`) filling
    `cells`, or the empty-state line at the range's top-left when there is
    nothing to plot -- a pie with no area is a rendering bug, not a data point
    (§4.2)."""
    min_col, min_row, _, _ = range_boundaries(cells)
    if group.total == 0:
        theme.empty_state(sheet, min_row, empty_text, column=min_col)
        return
    header_row, last_row = chart_data.block(
        name,
        ["Bucket", "Count"],
        [[BUCKET_LABELS[key], group.counts[key]] for key in BUCKET_KEYS],
    )
    chart = theme.pie_of(
        chart_data.sheet,
        title=title,
        header_row=header_row,
        last_row=last_row,
        colors=BUCKET_COLOR_ORDER,
        legend=legend,
        percent_labels=percent_labels,
    )
    theme.place(sheet, chart, cells)


# --------------------------------------------------------------------------
# Report
# --------------------------------------------------------------------------

REPORT_WIDTHS = {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 3}

# Fixed rows (§4.1). Every block is fixed-height: dollars are grouped by
# primary community (E14), so that block is at most five rows plus a
# footnote, and "By community" sits at BY_COMMUNITY_ROW (P8).
KPI_ROW = 6
STATUS_ROW = 10
ACTIVITY_ROW = 26
DOLLARS_ROW = 42
BY_COMMUNITY_ROW = 58

# Charts fill G:M beside the A:E blocks, 15 rows tall (about 8 cm).
CHART_ROWS = 15

# Under the dollars block (E14): primary attribution, stated on the sheet.
DOLLARS_FOOTNOTE = "Dollars count a work order under its first community."


def _chart_cells(top: int) -> str:
    return f"G{top}:M{top + CHART_ROWS - 1}"


def _community_money(
    rows: list[ReportRow],
) -> list[tuple[str, Decimal, Decimal]]:
    """Labor and materials dollars per primary community, biggest combined
    first (E14).

    `primary_community` -- a row's first membership -- rather than the raw
    `community` column (NULL on every imported row) or the full membership
    list (a row named in two communities must not have its dollars counted
    twice). Decimals throughout -- openpyxl writes them natively, and the
    report's money is never a float anywhere else either."""
    totals: dict[str, list[Decimal]] = {}
    for row in rows:
        key = primary_community(row)
        bucket = totals.setdefault(key, [Decimal("0.00"), Decimal("0.00")])
        bucket[0] += row.labor_total
        bucket[1] += row.materials_total
    return sorted(
        ((name, labor, materials) for name, (labor, materials) in totals.items()),
        key=lambda entry: (-(entry[1] + entry[2]), entry[0]),
    )


def _report_sheet(sheet: Worksheet, payload: DailyReport, chart_data: _ChartData) -> None:
    theme.setup_sheet(sheet, tab_color=theme.BRAND_RED)
    theme.set_widths(sheet, REPORT_WIDTHS)
    week = payload.week
    generated = payload.generated_at.astimezone(labor_day.CENTRAL)
    theme.title_block(
        sheet,
        "Weekly Work Order Report",
        [
            f"{payload.day.strftime('%a, %b %d, %Y')} · week of "
            f"{week.start.isoformat()} – {week.end.isoformat()} (week to date)",
            f"Generated {generated.strftime('%Y-%m-%d %H:%M')} Central",
            _population_caveat(payload),
        ],
    )
    company = payload.distribution.company
    sections = payload.sections

    # KPI strip. "Open" is the three live buckets; Closed is the week's
    # output -- the deleted pipeline chart's numbers, restated (§4.1).
    open_count = company.total - company.counts[BUCKET_CLOSED]
    tiles = [
        ("Open work orders", open_count),
        (BUCKET_LABELS["accepted"], company.counts["accepted"]),
        (BUCKET_LABELS["in_progress"], company.counts["in_progress"]),
        (BUCKET_LABELS["ready_to_close"], company.counts["ready_to_close"]),
        ("Closed this week", company.counts[BUCKET_CLOSED]),
    ]
    for column, (label, value) in enumerate(tiles, start=1):
        theme.kpi(sheet, KPI_ROW, column, label, value)

    # Company status: the four-slice pie beside its exact-value block (E9).
    theme.section(sheet, STATUS_ROW, "Company status")
    theme.header_row(sheet, STATUS_ROW + 1, ("Bucket", "Count", "Share"))
    theme.write_rows(
        sheet,
        STATUS_ROW + 2,
        _bucket_rows(company),
        formats={1: theme.COUNT, 2: theme.PERCENT},
    )
    _pie(
        sheet,
        chart_data,
        name="company",
        title="Company status",
        group=company,
        cells=_chart_cells(STATUS_ROW),
        legend="r",
        percent_labels=True,
    )

    # Activity: section counts, never row tallies -- `closing` can be capped
    # and the page follows the same rule (X8).
    theme.section(sheet, ACTIVITY_ROW, "Activity")
    activity = [
        ["Closed", sections.closed_today.count, sections.closed_week.count],
        ["New", sections.new_today.count, sections.new_week.count],
    ]
    theme.header_row(sheet, ACTIVITY_ROW + 1, (None, "Today", "Week to date"))
    theme.write_rows(
        sheet, ACTIVITY_ROW + 2, activity, formats={1: theme.COUNT, 2: theme.COUNT}
    )
    auto_today = sections.closed_today.auto_closed_count
    auto_week = sections.closed_week.auto_closed_count
    if auto_today or auto_week:
        # The sweep's closes are outside every figure on this sheet; say how
        # many, in the page's own phrasing, so the file reads like the screen.
        theme.note(
            sheet,
            ACTIVITY_ROW + 4,
            f"Closed today excludes {auto_today} closed in NetFacilities; "
            f"this week excludes {auto_week}.",
        )
    header_row, last_row = chart_data.block("activity", [None, "Today", "Week to date"], activity)
    theme.place(
        sheet,
        theme.column_chart_of(
            chart_data.sheet,
            title="Activity",
            header_row=header_row,
            last_row=last_row,
            y_title="Work orders",
        ),
        _chart_cells(ACTIVITY_ROW),
    )

    # Dollars closed this week, stacked labor over materials (X11).
    theme.section(sheet, DOLLARS_ROW, "Dollars closed this week")
    money = _community_money(sections.closed_week.rows) or [
        (PLACEHOLDER, Decimal("0.00"), Decimal("0.00"))
    ]
    theme.header_row(sheet, DOLLARS_ROW + 1, ("Community", "Labor", "Materials", "Total"))
    dollars_last = theme.write_rows(
        sheet,
        DOLLARS_ROW + 2,
        [[name, labor, materials, labor + materials] for name, labor, materials in money],
        formats={1: theme.MONEY, 2: theme.MONEY, 3: theme.MONEY},
    )
    theme.note(sheet, dollars_last + 1, DOLLARS_FOOTNOTE)
    header_row, last_row = chart_data.block(
        "dollars",
        ["Community", "Labor", "Materials"],
        [[name, labor, materials] for name, labor, materials in money],
    )
    theme.place(
        sheet,
        theme.column_chart_of(
            chart_data.sheet,
            title="Dollars closed this week",
            header_row=header_row,
            last_row=last_row,
            stacked=True,
            y_title="Dollars",
        ),
        _chart_cells(DOLLARS_ROW),
    )

    # By community: the four buckets per community, then the Ready-to-close
    # three-status split as a footnote (E11) and the memberships caveat.
    # Fixed row (P8): the dollars block above is bounded by E14.
    top = BY_COMMUNITY_ROW
    theme.section(sheet, top, "By community")
    last = theme.table_of(
        sheet,
        name="ByCommunity",
        row=top + 1,
        headers=("Community", *BUCKET_LABELS.values(), "Total"),
        rows=[
            [entry.label, *(entry.counts[key] for key in BUCKET_KEYS), entry.total]
            for entry in payload.distribution.communities
        ],
        formats={offset: theme.COUNT for offset in range(1, 6)},
    )
    by_status = sections.closing.by_status
    split = ", ".join(
        f"{by_status.get(status, 0):,} {STATUS_LABELS[status].lower()}"
        for status in CLOSING_STATUSES
    )
    theme.note(sheet, last + 2, f"Ready to close = {split}.")
    theme.note(
        sheet,
        last + 3,
        "A work order named in two communities is counted in both; community "
        "rows do not sum to the company figures above.",
    )


# --------------------------------------------------------------------------
# Community sheets
# --------------------------------------------------------------------------

# A = 22, B-J = 11 uniform, so the 3 x 3 grid below lines up (§4.2).
COMMUNITY_WIDTHS = {"A": 22, **{letter: 11 for letter in "BCDEFGHIJ"}}

COMMUNITY_PIE_CELLS = "A6:D20"
COMMUNITY_BLOCK_COLUMN = 6  # F6:H10 -- one column of air after the pie
GRID_SECTION_ROW = 22  # heading; the "Other" note sits on row 23
DETAIL_ROW = 64

# Nine cards, row-major: B/E/H, thirteen rows apart, each 12 rows tall.
GRID_CELLS: tuple[str, ...] = tuple(
    f"{first}{top}:{last}{top + 12}"
    for top in (24, 37, 50)
    for first, last in (("B", "D"), ("E", "G"), ("H", "J"))
)


def _community_sheet(
    sheet: Worksheet,
    payload: DailyReport,
    community: CommunityDistribution,
    chart_data: _ChartData,
) -> None:
    theme.setup_sheet(sheet, tab_color=theme.MUTED)
    theme.set_widths(sheet, COMMUNITY_WIDTHS)
    theme.title_block(
        sheet,
        community.label,
        [
            f"Weekly status · {community.total:,} work orders",
            _population_caveat(payload),
        ],
    )
    if community.total == 0:
        # No charts of zeros (§4.2): the line stands in for the whole body.
        theme.empty_state(sheet, 5, EMPTY_COMMUNITY_STATE)
        return

    theme.section(sheet, 5, "Status", span=10)
    _pie(
        sheet,
        chart_data,
        name=community.key,
        title=f"{community.label} · status",
        group=community,
        cells=COMMUNITY_PIE_CELLS,
        legend="r",
        percent_labels=True,
        empty_text=EMPTY_COMMUNITY_STATE,
    )
    theme.header_row(sheet, 6, ("Bucket", "Count", "Share"), column=COMMUNITY_BLOCK_COLUMN)
    theme.write_rows(
        sheet,
        7,
        _bucket_rows(community),
        column=COMMUNITY_BLOCK_COLUMN,
        formats={1: theme.COUNT, 2: theme.PERCENT},
    )

    theme.section(sheet, GRID_SECTION_ROW, "By service type", span=10)
    cards, folded = grid_of(community.service_types)
    if folded:
        theme.note(
            sheet,
            GRID_SECTION_ROW + 1,
            f'"Other" combines {folded} further service types.',
        )
    for card, cells in zip(cards, GRID_CELLS):
        _pie(
            sheet,
            chart_data,
            name=f"{community.key}:{card.key}",
            title=f"{card.label} · {card.total:,}",
            group=card,
            cells=cells,
            legend=None,  # E15: the status pie above is the shared key
            percent_labels=False,
        )

    theme.section(sheet, DETAIL_ROW, "Service type detail", span=10)
    theme.table_of(
        sheet,
        name=f"ServiceTypes_{community.key}",
        row=DETAIL_ROW + 1,
        headers=("Service type", *BUCKET_LABELS.values(), "Total"),
        rows=[
            [entry.label, *(entry.counts[key] for key in BUCKET_KEYS), entry.total]
            for entry in community.service_types
        ],
        formats={offset: theme.COUNT for offset in range(1, 6)},
    )
