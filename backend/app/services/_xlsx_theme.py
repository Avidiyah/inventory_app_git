"""House style for the daily report workbook (redesign spec §4.0, E12).

Layer: services (pure -- openpyxl objects in, openpyxl objects out). Fonts,
fills, number formats, the bucket palette, table and chart construction, so
`work_order_report_xlsx.py` is sheet composition and nothing else.

Two things about this module are load-bearing:

**Charts are placed by cell range, not by centimetres.** openpyxl's default
one-cell anchor sizes a chart in absolute units, so a grid of pies lines up
only if column widths happen to match the font's pixel metrics -- which
change with whatever font Excel substitutes. `place()` uses a two-cell anchor
instead: the chart fills exactly the cells named, whatever the widths.

**Slice colors are set per data point, in bucket order, on every pie.** Excel
would otherwise assign its theme's accent sequence and Accepted would be a
different color on every sheet. `pie_of` takes the colors explicitly so the
palette below is the only place they exist.
"""

from __future__ import annotations

import math
from typing import Optional, Sequence

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# --- palette (§4.0) ---------------------------------------------------------

BRAND_RED = "C8102E"  # titles and the KPI rule: the brand, never a slice
INK = "1C1D20"
MUTED = "5A5C60"
RULE = "D8D9DB"
TAB_GRAY = "B7B9BC"
WHITE = "FFFFFF"

# Bucket key -> fill. Traceable to the `--wo-status-*` tokens the Graphs tab
# uses, darkened where the token was tuned for a dark canvas and would wash
# out on white paper.
BUCKET_COLORS: dict[str, str] = {
    "accepted": "9CA3AF",  # --wo-status-created #D1D5DB, darkened
    "in_progress": "D97706",  # --wo-status-in-progress #FACC15, darkened
    "ready_to_close": "6D28D9",  # --wo-status-ready-to-complete, as-is
    "closed": "15803D",  # --wo-status-review, as-is
}

# The two-series treatment for the Activity and Dollars charts. Not the
# bucket palette: those are not bucket charts and must not borrow it (§4.1).
SERIES_COLORS: tuple[str, str] = (BRAND_RED, MUTED)

# --- type (§4.0) ------------------------------------------------------------

FONT_NAME = "Aptos Narrow"  # Excel falls back to Calibri where it is missing


def font(**overrides) -> Font:
    return Font(**{"name": FONT_NAME, "size": 10, **overrides})


BODY = font()
TITLE = font(size=18, bold=True, color=BRAND_RED)
SUBTITLE = font(size=9, italic=True, color=MUTED)
SECTION = font(size=11, bold=True, color=INK)
HEADER = font(bold=True, color=WHITE)
KPI_LABEL = font(size=9, color=MUTED)
KPI_VALUE = font(size=20, bold=True, color=INK)

HEADER_FILL = PatternFill("solid", fgColor=INK)
SECTION_RULE = Border(bottom=Side(style="thin", color=RULE))
KPI_RULE = Border(bottom=Side(style="medium", color=BRAND_RED))
TOP = Alignment(vertical="top")
TOP_WRAPPED = Alignment(vertical="top", wrap_text=True)

COUNT = "#,##0"
MONEY = "$#,##0.00"
PERCENT = "0.0%"
DATE = "yyyy-mm-dd hh:mm"

TABLE_STYLE = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)

# The Notes column (§4.3): 60 characters wide, wrapped, at most four lines
# tall. 13.5pt is one line of 10pt text with Excel's default leading.
NOTES_WIDTH = 60
NOTES_MAX_LINES = 4
LINE_POINTS = 13.5


# --- sheet furniture --------------------------------------------------------


def setup_sheet(
    sheet: Worksheet,
    *,
    tab_color: str,
    freeze: Optional[str] = "A5",
    gridlines: bool = False,
    print_title_rows: Optional[str] = None,
) -> None:
    """Tab color, gridlines, freeze panes, and the print setup every sheet
    shares: landscape, one page wide, half-inch margins."""
    sheet.sheet_properties.tabColor = tab_color
    sheet.sheet_view.showGridLines = gridlines
    sheet.freeze_panes = freeze
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    for side in ("left", "right", "top", "bottom"):
        setattr(sheet.page_margins, side, 0.5)
    if print_title_rows:
        sheet.print_title_rows = print_title_rows


def set_widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def title_block(sheet: Worksheet, title: str, lines: Sequence[str]) -> None:
    """Rows 1-4: the title, then up to three subtitle / caveat lines."""
    sheet.cell(row=1, column=1, value=title).font = TITLE
    for offset, line in enumerate(lines, start=2):
        sheet.cell(row=offset, column=1, value=line).font = SUBTITLE


def section(sheet: Worksheet, row: int, text: str, *, span: int = 6) -> None:
    """A section heading with a 1pt rule under it across `span` columns."""
    sheet.cell(row=row, column=1, value=text).font = SECTION
    for column in range(1, span + 1):
        sheet.cell(row=row, column=column).border = SECTION_RULE


def note(sheet: Worksheet, row: int, text: str, *, column: int = 1) -> None:
    sheet.cell(row=row, column=column, value=text).font = SUBTITLE


def empty_state(sheet: Worksheet, row: int, text: str, *, column: int = 1) -> None:
    sheet.cell(row=row, column=column, value=text).font = font(italic=True, color=MUTED)


def kpi(sheet: Worksheet, row: int, column: int, label: str, value) -> None:
    """One KPI tile: label above value, brand-red rule under the value."""
    sheet.cell(row=row, column=column, value=label).font = KPI_LABEL
    cell = sheet.cell(row=row + 1, column=column, value=value)
    cell.font = KPI_VALUE
    cell.number_format = COUNT
    cell.border = KPI_RULE


def header_row(
    sheet: Worksheet, row: int, headers: Sequence[str], *, column: int = 1
) -> None:
    """White-on-ink header cells: plain blocks and Excel Table headers alike
    (the table style leaves explicit fills alone)."""
    for offset, header in enumerate(headers):
        cell = sheet.cell(row=row, column=column + offset, value=header)
        cell.font = HEADER
        cell.fill = HEADER_FILL


def write_rows(
    sheet: Worksheet,
    row: int,
    rows: Sequence[Sequence],
    *,
    column: int = 1,
    formats: Optional[dict[int, str]] = None,
    alignment: Optional[Alignment] = None,
) -> int:
    """Write `rows` from `row` down. `formats` maps a 0-based column offset to
    a number format. Returns the last row written (`row - 1` if none)."""
    formats = formats or {}
    for index, values in enumerate(rows):
        for offset, value in enumerate(values):
            cell = sheet.cell(row=row + index, column=column + offset, value=value)
            cell.font = BODY
            if offset in formats:
                cell.number_format = formats[offset]
            if alignment is not None:
                cell.alignment = alignment
    return row + len(rows) - 1


def table_of(
    sheet: Worksheet,
    *,
    name: str,
    row: int,
    headers: Sequence[str],
    rows: Sequence[Sequence],
    column: int = 1,
    formats: Optional[dict[int, str]] = None,
    alignment: Optional[Alignment] = None,
) -> int:
    """A banded Excel Table with autofilter, header at `row`. Returns the last
    row written.

    With no data rows the header is still written but no Table is created:
    Excel treats a header-only table as a file to repair. `name` must be
    unique in the workbook and contain no spaces."""
    header_row(sheet, row, headers, column=column)
    last = write_rows(
        sheet, row + 1, rows, column=column, formats=formats, alignment=alignment
    )
    if rows:
        first_letter = sheet.cell(row=row, column=column).column_letter
        last_letter = sheet.cell(row=row, column=column + len(headers) - 1).column_letter
        table = Table(displayName=name, ref=f"{first_letter}{row}:{last_letter}{last}")
        table.tableStyleInfo = TABLE_STYLE
        sheet.add_table(table)
    return last


def notes_row_height(text: Optional[str]) -> Optional[float]:
    """Row height for a wrapped Notes cell: one line per `NOTES_WIDTH`
    characters, capped at `NOTES_MAX_LINES`. `None` means leave the default."""
    if not text:
        return None
    lines = sum(
        max(1, math.ceil(len(part) / NOTES_WIDTH)) for part in text.splitlines()
    )
    if lines <= 1:
        return None
    return min(lines, NOTES_MAX_LINES) * LINE_POINTS


# --- charts -----------------------------------------------------------------


def pie_of(
    source: Worksheet,
    *,
    title: str,
    header_row: int,
    last_row: int,
    colors: Sequence[str],
    legend: Optional[str] = "r",
    percent_labels: bool = True,
) -> PieChart:
    """A pie over a two-column block on `source`: labels in column A, counts
    in column B, `header_row` holding the series title.

    `colors` are applied per data point in row order -- pass the bucket
    palette in bucket order. `firstSliceAng = 0` keeps the first slice at 12
    o'clock; `visible_cells_only = False` lets Excel plot a block that lives
    on a hidden sheet (E7)."""
    chart = PieChart()
    chart.title = title
    chart.firstSliceAng = 0
    chart.visible_cells_only = False
    chart.add_data(
        Reference(source, min_col=2, min_row=header_row, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(source, min_col=1, min_row=header_row + 1, max_row=last_row)
    )
    series = chart.series[0]
    for index, color in enumerate(colors):
        point = DataPoint(idx=index)
        point.graphicalProperties = GraphicalProperties(solidFill=color)
        series.dPt.append(point)
    if percent_labels:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLeaderLines = False
    if legend is None:
        chart.legend = None
    else:
        chart.legend.position = legend
    return chart


def column_chart_of(
    source: Worksheet,
    *,
    title: str,
    header_row: int,
    last_row: int,
    series_columns: tuple[int, int] = (2, 3),
    stacked: bool = False,
    y_title: Optional[str] = None,
) -> BarChart:
    """A vertical column chart over a block on `source`: categories in column
    A, one series per column in `series_columns`, titles from `header_row`.
    Brand red then neutral -- never the bucket palette."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.visible_cells_only = False
    # openpyxl 3.1 leaves `<c:delete>` unwritten, which current Excel reads
    # as "axis deleted": say so explicitly or the chart renders without axes.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100
    if y_title:
        chart.y_axis.title = y_title
    first, last = series_columns
    chart.add_data(
        Reference(
            source, min_col=first, max_col=last, min_row=header_row, max_row=last_row
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(source, min_col=1, min_row=header_row + 1, max_row=last_row)
    )
    for series, color in zip(chart.series, SERIES_COLORS):
        series.graphicalProperties = GraphicalProperties(solidFill=color)
    return chart


def place(sheet: Worksheet, chart, cells: str) -> None:
    """Anchor `chart` so it fills exactly `cells` -- an inclusive A1 range such
    as `"B24:D36"` -- whatever the column widths and row heights are."""
    min_col, min_row, max_col, max_row = range_boundaries(cells)
    chart.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=min_col - 1, row=min_row - 1),
        to=AnchorMarker(col=max_col, row=max_row),
    )
    sheet.add_chart(chart)
