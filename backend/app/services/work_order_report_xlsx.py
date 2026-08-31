"""The Admin daily report as an Excel workbook.

Layer: services. The third renderer of `work_order_report.daily_report`'s
payload, beside the JSON route and `report_csv` -- a pure function of that
payload, no queries and no clock, so the file and the screen cannot disagree
(parent spec R9 / X3, redesign E13). Styling lives in `_xlsx_theme.py` (E12);
this module is sheet composition only.

Spec: docs/superpowers/specs/2026-08-30-hub-report-xlsx-redesign-design.md

Sheet order is the reading order: `Report`, one sheet per community, `Work
Orders`, then the machine sheets -- hidden `Chart Data`, and `Data` last.

Three things about this module are load-bearing:

**openpyxl discards charts across a load/save cycle.** So there is no
committed `.xlsx` template -- the workbook is built in code, and the charts
cannot be read back in tests: `tests/test_work_order_report_xlsx.py` asserts
them over the saved bytes with `zipfile`.

**Every chart reads the hidden `Chart Data` sheet (E7)**, one labelled block
per chart written by cursor, with `visible_cells_only = False` so Excel plots
it. The designed sheets carry the same numbers as styled tables, from the
same payload -- never as chart sources.

**`Data` is `report_csv`, cell for cell (E10 / X5)**, money-as-text included,
so save-as-CSV from Excel still round-trips through `parse_import_row`.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.domain import labor_day
from app.domain import work_orders as wo
from app.services import _xlsx_theme as theme
from app.services.work_order_report import (
    CLOSING_STATUSES,
    CSV_SECTION_HEADER,
    SECTION_ORDER,
    STATUS_LABELS,
    DailyReport,
    ReportRow,
)
from app.services.work_order_report_buckets import (
    BUCKET_CLOSED,
    BUCKET_KEYS,
    BUCKET_LABELS,
    CommunityDistribution,
    communities_of,
    grid_of,
    primary_community,
    row_bucket,
)

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

CHART_DATA_SHEET = "Chart Data"

# Physical order. Excel opens on the first; `Chart Data` is hidden, so the
# tab strip shows eight and ends on `Data` (E6, plan P5).
SHEET_NAMES: tuple[str, ...] = (
    "Report",
    *(wo.COMMUNITY_LABELS[key] for key in wo.ALL_COMMUNITY_FILTERS),
    "Work Orders",
    CHART_DATA_SHEET,
    "Data",
)

PLACEHOLDER = "(none)"
EMPTY_STATE = "No live or recently closed work orders."
EMPTY_COMMUNITY_STATE = "No live or recently closed work orders in this community."

BUCKET_COLOR_ORDER = [theme.BUCKET_COLORS[key] for key in BUCKET_KEYS]


def report_xlsx(payload: DailyReport) -> bytes:
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "Report"
    community_sheets = {
        key: workbook.create_sheet(wo.COMMUNITY_LABELS[key])
        for key in wo.ALL_COMMUNITY_FILTERS
    }
    work_orders_sheet = workbook.create_sheet("Work Orders")
    chart_data = _ChartData(workbook.create_sheet(CHART_DATA_SHEET))
    chart_data.sheet.sheet_state = "hidden"
    data_sheet = workbook.create_sheet("Data")

    _report_sheet(report_sheet, payload, chart_data)
    for community in payload.distribution.communities:
        _community_sheet(community_sheets[community.key], payload, community, chart_data)
    _work_orders_sheet(work_orders_sheet, payload)
    _data_sheet(data_sheet, payload)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_xlsx_filename(payload: DailyReport) -> str:
    """Named for the period it covers, not the moment of export -- the same
    timesheet convention `report_filename` follows (user-hub-design.md D14)."""
    return f"wo-report_{payload.day.isoformat()}.xlsx"


# --------------------------------------------------------------------------
# Shared pieces
# --------------------------------------------------------------------------


class _ChartData:
    """Cursor over the hidden source sheet: one named block per chart (E7).

    Each block is a one-cell name, a header row, then its rows, then a blank
    row -- so a person who unhides the sheet can find what a chart reads."""

    def __init__(self, sheet: Worksheet) -> None:
        self.sheet = sheet
        self.row = 1

    def block(self, name: str, header: list, rows: list[list]) -> tuple[int, int]:
        """Write the block; return `(header_row, last_row)` for a `Reference`."""
        self.sheet.cell(row=self.row, column=1, value=name)
        header_row = self.row + 1
        for offset, value in enumerate(header, start=1):
            self.sheet.cell(row=header_row, column=offset, value=value)
        for index, values in enumerate(rows, start=1):
            for offset, value in enumerate(values, start=1):
                self.sheet.cell(row=header_row + index, column=offset, value=value)
        last_row = header_row + len(rows)
        self.row = last_row + 2
        return header_row, last_row


def _population_caveat(payload: DailyReport) -> str:
    """§2.1, on every designed sheet: what the population is, and why the
    community totals do not sum to the company total."""
    generated = payload.generated_at.astimezone(labor_day.CENTRAL)
    week = payload.week
    return (
        f"Live work orders as of {generated:%Y-%m-%d %H:%M} Central, plus work "
        f"orders closed {week.start.isoformat()} – {week.end.isoformat()}. A work "
        "order named in two communities is counted in both; community totals do "
        "not sum to the company total."
    )


def _bucket_rows(group: CommunityDistribution) -> list[list]:
    """Label / count / share, in bucket order. Shares are real fractions."""
    return [
        [
            BUCKET_LABELS[key],
            group.counts[key],
            (group.counts[key] / group.total) if group.total else 0,
        ]
        for key in BUCKET_KEYS
    ]


def _pie(
    sheet: Worksheet,
    chart_data: _ChartData,
    *,
    name: str,
    title: str,
    group,
    cells: str,
    legend: Optional[str],
    percent_labels: bool,
    empty_text: str = EMPTY_STATE,
) -> None:
    """A bucket pie for `group` (anything with `.total` and `.counts`) filling
    `cells`, or the empty-state line at the range's top-left when there is
    nothing to plot -- a pie with no area is a rendering bug, not a data point
    (§4.2)."""
    min_col, min_row, _, _ = range_boundaries(cells)
    if group.total == 0:
        theme.empty_state(sheet, min_row, empty_text, column=min_col)
        return
    header_row, last_row = chart_data.block(
        name,
        ["Bucket", "Count"],
        [[BUCKET_LABELS[key], group.counts[key]] for key in BUCKET_KEYS],
    )
    chart = theme.pie_of(
        chart_data.sheet,
        title=title,
        header_row=header_row,
        last_row=last_row,
        colors=BUCKET_COLOR_ORDER,
        legend=legend,
        percent_labels=percent_labels,
    )
    theme.place(sheet, chart, cells)


# --------------------------------------------------------------------------
# Report
# --------------------------------------------------------------------------

REPORT_WIDTHS = {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 3}

# Fixed rows (§4.1). Every block is fixed-height: dollars are grouped by
# primary community (E14), so that block is at most five rows plus a
# footnote, and "By community" sits at BY_COMMUNITY_ROW (P8).
KPI_ROW = 6
STATUS_ROW = 10
ACTIVITY_ROW = 26
DOLLARS_ROW = 42
BY_COMMUNITY_ROW = 58

# Charts fill G:M beside the A:E blocks, 15 rows tall (about 8 cm).
CHART_ROWS = 15

# Under the dollars block (E14): primary attribution, stated on the sheet.
DOLLARS_FOOTNOTE = "Dollars count a work order under its first community."


def _chart_cells(top: int) -> str:
    return f"G{top}:M{top + CHART_ROWS - 1}"


def _community_money(
    rows: list[ReportRow],
) -> list[tuple[str, Decimal, Decimal]]:
    """Labor and materials dollars per primary community, biggest combined
    first (E14).

    `primary_community` -- a row's first membership -- rather than the raw
    `community` column (NULL on every imported row) or the full membership
    list (a row named in two communities must not have its dollars counted
    twice). Decimals throughout -- openpyxl writes them natively, and the
    report's money is never a float anywhere else either."""
    totals: dict[str, list[Decimal]] = {}
    for row in rows:
        key = primary_community(row)
        bucket = totals.setdefault(key, [Decimal("0.00"), Decimal("0.00")])
        bucket[0] += row.labor_total
        bucket[1] += row.materials_total
    return sorted(
        ((name, labor, materials) for name, (labor, materials) in totals.items()),
        key=lambda entry: (-(entry[1] + entry[2]), entry[0]),
    )


def _report_sheet(sheet: Worksheet, payload: DailyReport, chart_data: _ChartData) -> None:
    theme.setup_sheet(sheet, tab_color=theme.BRAND_RED)
    theme.set_widths(sheet, REPORT_WIDTHS)
    week = payload.week
    generated = payload.generated_at.astimezone(labor_day.CENTRAL)
    theme.title_block(
        sheet,
        "Weekly Work Order Report",
        [
            f"{payload.day.strftime('%a, %b %d, %Y')} · week of "
            f"{week.start.isoformat()} – {week.end.isoformat()} (week to date)",
            f"Generated {generated.strftime('%Y-%m-%d %H:%M')} Central",
            _population_caveat(payload),
        ],
    )
    company = payload.distribution.company
    sections = payload.sections

    # KPI strip. "Open" is the three live buckets; Closed is the week's
    # output -- the deleted pipeline chart's numbers, restated (§4.1).
    open_count = company.total - company.counts[BUCKET_CLOSED]
    tiles = [
        ("Open work orders", open_count),
        (BUCKET_LABELS["accepted"], company.counts["accepted"]),
        (BUCKET_LABELS["in_progress"], company.counts["in_progress"]),
        (BUCKET_LABELS["ready_to_close"], company.counts["ready_to_close"]),
        ("Closed this week", company.counts[BUCKET_CLOSED]),
    ]
    for column, (label, value) in enumerate(tiles, start=1):
        theme.kpi(sheet, KPI_ROW, column, label, value)

    # Company status: the four-slice pie beside its exact-value block (E9).
    theme.section(sheet, STATUS_ROW, "Company status")
    theme.header_row(sheet, STATUS_ROW + 1, ("Bucket", "Count", "Share"))
    theme.write_rows(
        sheet,
        STATUS_ROW + 2,
        _bucket_rows(company),
        formats={1: theme.COUNT, 2: theme.PERCENT},
    )
    _pie(
        sheet,
        chart_data,
        name="company",
        title="Company status",
        group=company,
        cells=_chart_cells(STATUS_ROW),
        legend="r",
        percent_labels=True,
    )

    # Activity: section counts, never row tallies -- `closing` can be capped
    # and the page follows the same rule (X8).
    theme.section(sheet, ACTIVITY_ROW, "Activity")
    activity = [
        ["Closed", sections.closed_today.count, sections.closed_week.count],
        ["New", sections.new_today.count, sections.new_week.count],
    ]
    theme.header_row(sheet, ACTIVITY_ROW + 1, (None, "Today", "Week to date"))
    theme.write_rows(
        sheet, ACTIVITY_ROW + 2, activity, formats={1: theme.COUNT, 2: theme.COUNT}
    )
    auto_today = sections.closed_today.auto_closed_count
    auto_week = sections.closed_week.auto_closed_count
    if auto_today or auto_week:
        # The page's own phrasing (R10), so the file reads like the screen.
        theme.note(
            sheet,
            ACTIVITY_ROW + 4,
            f"Closed today includes ({auto_today} in NetFacilities); "
            f"this week ({auto_week} in NetFacilities)",
        )
    header_row, last_row = chart_data.block("activity", [None, "Today", "Week to date"], activity)
    theme.place(
        sheet,
        theme.column_chart_of(
            chart_data.sheet,
            title="Activity",
            header_row=header_row,
            last_row=last_row,
            y_title="Work orders",
        ),
        _chart_cells(ACTIVITY_ROW),
    )

    # Dollars closed this week, stacked labor over materials (X11).
    theme.section(sheet, DOLLARS_ROW, "Dollars closed this week")
    money = _community_money(sections.closed_week.rows) or [
        (PLACEHOLDER, Decimal("0.00"), Decimal("0.00"))
    ]
    theme.header_row(sheet, DOLLARS_ROW + 1, ("Community", "Labor", "Materials", "Total"))
    dollars_last = theme.write_rows(
        sheet,
        DOLLARS_ROW + 2,
        [[name, labor, materials, labor + materials] for name, labor, materials in money],
        formats={1: theme.MONEY, 2: theme.MONEY, 3: theme.MONEY},
    )
    theme.note(sheet, dollars_last + 1, DOLLARS_FOOTNOTE)
    header_row, last_row = chart_data.block(
        "dollars",
        ["Community", "Labor", "Materials"],
        [[name, labor, materials] for name, labor, materials in money],
    )
    theme.place(
        sheet,
        theme.column_chart_of(
            chart_data.sheet,
            title="Dollars closed this week",
            header_row=header_row,
            last_row=last_row,
            stacked=True,
            y_title="Dollars",
        ),
        _chart_cells(DOLLARS_ROW),
    )

    # By community: the four buckets per community, then the Ready-to-close
    # three-status split as a footnote (E11) and the memberships caveat.
    # Fixed row (P8): the dollars block above is bounded by E14.
    top = BY_COMMUNITY_ROW
    theme.section(sheet, top, "By community")
    last = theme.table_of(
        sheet,
        name="ByCommunity",
        row=top + 1,
        headers=("Community", *BUCKET_LABELS.values(), "Total"),
        rows=[
            [entry.label, *(entry.counts[key] for key in BUCKET_KEYS), entry.total]
            for entry in payload.distribution.communities
        ],
        formats={offset: theme.COUNT for offset in range(1, 6)},
    )
    by_status = sections.closing.by_status
    split = ", ".join(
        f"{by_status.get(status, 0):,} {STATUS_LABELS[status].lower()}"
        for status in CLOSING_STATUSES
    )
    theme.note(sheet, last + 2, f"Ready to close = {split}.")
    theme.note(
        sheet,
        last + 3,
        "A work order named in two communities is counted in both; community "
        "rows do not sum to the company figures above.",
    )


# --------------------------------------------------------------------------
# Community sheets  (body added in Task 6)
# --------------------------------------------------------------------------


def _community_sheet(
    sheet: Worksheet,
    payload: DailyReport,
    community: CommunityDistribution,
    chart_data: _ChartData,
) -> None:
    theme.setup_sheet(sheet, tab_color=theme.MUTED)
    theme.title_block(
        sheet,
        community.label,
        [
            f"Weekly status · {community.total:,} work orders",
            _population_caveat(payload),
        ],
    )


# --------------------------------------------------------------------------
# Work Orders
# --------------------------------------------------------------------------

WORK_ORDERS_HEADER_ROW = 5

WORK_ORDER_HEADERS: tuple[str, ...] = (
    "WORK ORDER",
    "LOCATION",
    "NOTES",
    "BUCKET",
    "STATUS",
    "COMMUNITIES",
    "SERVICE TYPE",
    "PRIORITY",
    "BUILDING",
    "UNIT",
    "SUPERVISOR",
    "TECHNICIANS",
    "MATERIAL LINES",
    "MATERIALS TOTAL",
    "LABOR MINUTES",
    "LABOR TOTAL",
    "TOTAL",
    "CREATED AT",
    "COMPLETED AT",
    "CLOSED AT",
    "SECTIONS",
)

WORK_ORDER_WIDTHS: tuple[int, ...] = (
    14, 34, theme.NOTES_WIDTH, 14, 16, 22, 18, 12, 10, 10, 18, 24, 12, 14, 12, 14, 14, 18, 18, 18, 20,
)

# 0-based column offset -> number format.
WORK_ORDER_FORMATS: dict[int, str] = {
    12: theme.COUNT,
    13: theme.MONEY,
    14: theme.COUNT,
    15: theme.MONEY,
    16: theme.MONEY,
    17: theme.DATE,
    18: theme.DATE,
    19: theme.DATE,
}


def _excel_utc(value: Optional[datetime]) -> Optional[datetime]:
    """UTC, naive: openpyxl refuses tz-aware datetimes. UTC on purpose -- the
    same seam `export_row` writes (§5); the covered period is in the title."""
    if value is None:
        return None
    return labor_day.as_utc(value).astimezone(timezone.utc).replace(tzinfo=None)


def _sections_of(payload: DailyReport) -> dict[UUID, list[str]]:
    """Which report sections each work order appeared in, in SECTION_ORDER --
    the `Data` sheet's SECTION filter folded into one deduped row."""
    seen: dict[UUID, list[str]] = {}
    for key in SECTION_ORDER:
        for row in getattr(payload.sections, key).rows:
            seen.setdefault(row.work_order_id, []).append(key)
    return seen


def _work_order_cells(row: ReportRow, sections: dict[UUID, list[str]]) -> list:
    return [
        row.number,
        row.location,
        row.notes,
        BUCKET_LABELS[row_bucket(row)],
        STATUS_LABELS.get(row.status, row.status),
        "; ".join(communities_of(row)),
        row.service_type,
        row.priority,
        row.building_number,
        row.unit_number,
        row.supervisor_name,
        "; ".join(row.technician_names),
        row.material_lines,
        row.materials_total,
        row.labor_minutes,
        row.labor_total,
        row.total,
        _excel_utc(row.created_at),
        _excel_utc(row.completed_at),
        _excel_utc(row.archived_at),
        "; ".join(sections.get(row.work_order_id, [])),
    ]


def _work_orders_sheet(sheet: Worksheet, payload: DailyReport) -> None:
    """One row per work order over the E1 population, deduped, in reading
    order (§4.3). Money is numeric here -- this sheet is for reading and
    pivoting, and it is not the re-import path."""
    theme.setup_sheet(
        sheet,
        tab_color=theme.INK,
        freeze=f"D{WORK_ORDERS_HEADER_ROW + 1}",
        print_title_rows=f"{WORK_ORDERS_HEADER_ROW}:{WORK_ORDERS_HEADER_ROW}",
    )
    theme.set_widths(
        sheet,
        {get_column_letter(index): width for index, width in enumerate(WORK_ORDER_WIDTHS, start=1)},
    )
    week = payload.week
    theme.title_block(
        sheet,
        "Work Orders",
        [
            f"{len(payload.all_rows):,} work orders · live now plus closed "
            f"{week.start.isoformat()} – {week.end.isoformat()}",
            "One row per work order. Timestamps are UTC; the covered period is "
            "in the line above and in the filename.",
        ],
    )

    sections = _sections_of(payload)
    rows = [_work_order_cells(row, sections) for row in payload.all_rows]
    theme.table_of(
        sheet,
        name="WorkOrders",
        row=WORK_ORDERS_HEADER_ROW,
        headers=WORK_ORDER_HEADERS,
        rows=rows,
        formats=WORK_ORDER_FORMATS,
        alignment=theme.TOP,
    )
    first = WORK_ORDERS_HEADER_ROW + 1
    for index, row in enumerate(payload.all_rows, start=first):
        sheet.cell(row=index, column=3).alignment = theme.TOP_WRAPPED
        height = theme.notes_row_height(row.notes)
        if height is not None:
            sheet.row_dimensions[index].height = height
    if not rows:
        theme.empty_state(sheet, first, EMPTY_STATE)


# --------------------------------------------------------------------------
# Data
# --------------------------------------------------------------------------


def _data_sheet(sheet: Worksheet, payload: DailyReport) -> None:
    """`report_csv` as cells: same header, same section order, same values.

    Deliberately not coerced to numbers. The money columns stay the strings
    `export_row` produced, so Excel shows its "number stored as text" hint on
    them -- that is the price of the CSV being byte-identical, and the charts
    are immune because they read `Chart Data` instead."""
    theme.setup_sheet(sheet, tab_color=theme.TAB_GRAY, freeze="A2", gridlines=True)
    sheet.append([CSV_SECTION_HEADER, *wo.EXPORT_HEADERS])
    for cell in sheet[1]:
        cell.font = theme.font(bold=True)
    for key in SECTION_ORDER:
        # All five sections, so a row appears under both `closed_today` and
        # `closed_week` -- the CSV's filter-on-SECTION property, preserved.
        for row in getattr(payload.sections, key).rows:
            sheet.append([key, *row.export_cells])
