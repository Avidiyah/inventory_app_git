"""The daily report's Excel render.

`report_xlsx` is a pure function of a `DailyReport` (spec X3 / E13), so these
tests hand-build frozen payloads instead of touching the database -- no `db`
fixture, no dev-Postgres fencing. Charts cannot be read back through openpyxl,
so they are asserted over the saved bytes with `zipfile`.

Spec: docs/superpowers/specs/2026-08-30-hub-report-xlsx-redesign-design.md
"""

import csv
import io
import os
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date, datetime, timezone
from decimal import Decimal
from uuid import uuid4

import openpyxl
import pytest

from app.domain import work_orders as wo
from app.services import work_order_report as report
from app.services import work_order_report_buckets as buckets
from app.services import work_order_report_xlsx as xlsx

CLOSED_AT = datetime(2026, 8, 25, 15, 30, tzinfo=timezone.utc)


def _row(
    *,
    number="WO-1",
    status=wo.STATUS_CREATED,
    community=None,
    location="Bldg A",
    service_type=None,
    labor_total="0.00",
    materials_total="0.00",
    archived=False,
    notes=None,
    priority=None,
):
    """A `ReportRow` whose `export_cells` are shaped like `export_row`'s: 26
    values, three of them genuine ints, money as fixed-point strings. Live by
    default; `archived=True` makes it a row closed this week."""
    labor = Decimal(labor_total)
    materials = Decimal(materials_total)
    archived_at = CLOSED_AT if archived else None
    cells = [
        number,
        location or "",
        "",
        "",
        service_type or "",
        "",
        "Leaky faucet",
        status,
        "Tech One; Tech Two",
        "Sue",
        community or "",
        "3",
        "12",
        "manual",
        2,
        f"{materials:.2f}",
        60,
        60,
        f"{labor:.2f}",
        f"{materials + labor:.2f}",
        notes or "",
        "2026-08-25 14:00",
        "2026-08-25 15:00",
        "2026-08-25 15:00",
        "2026-08-25 15:30" if archived else "",
    ]
    assert len(cells) == len(wo.EXPORT_HEADERS)
    return report.ReportRow(
        work_order_id=uuid4(),
        number=number,
        status=status,
        community=community,
        location=location,
        building_number="3",
        unit_number="12",
        service_type=service_type,
        priority=priority,
        supervisor_name="Sue",
        technician_names=["Tech One", "Tech Two"],
        materials_total=materials,
        labor_minutes=60,
        labor_total=labor,
        total=materials + labor,
        created_at=datetime(2026, 8, 25, 14, 0, tzinfo=timezone.utc),
        completed_at=datetime(2026, 8, 25, 15, 0, tzinfo=timezone.utc),
        archived_at=archived_at,
        auto_closed=False,
        legacy=False,
        notes=notes,
        material_lines=2,
        export_cells=cells,
    )


def _payload(
    *,
    live=(),
    closed_today=(),
    closed_week=(),
    closing=(),
    new_today=(),
    new_week=(),
    auto_closed_today=0,
    by_status=None,
    closing_count=None,
    truncated=False,
):
    """The E1 population is `live` + `closed_week`; the sections are passed
    separately, exactly as `daily_report` composes them."""
    closing_rows = list(closing)
    all_rows = sorted([*live, *closed_week], key=report.reading_order)
    return report.DailyReport(
        generated_at=datetime(2026, 8, 25, 22, 30, tzinfo=timezone.utc),
        day=date(2026, 8, 25),
        week=report.ReportWeek(start=date(2026, 8, 24), end=date(2026, 8, 30)),
        sections=report.ReportSections(
            closed_today=report.ClosedSection(
                count=len(closed_today),
                auto_closed_count=auto_closed_today,
                rows=list(closed_today),
            ),
            closed_week=report.ClosedSection(
                count=len(closed_week), auto_closed_count=0, rows=list(closed_week)
            ),
            closing=report.ClosingSection(
                count=(closing_count if closing_count is not None else len(closing_rows)),
                by_status=(by_status if by_status is not None else {}),
                truncated=truncated,
                rows=closing_rows,
            ),
            new_today=report.NewSection(count=len(new_today), rows=list(new_today)),
            new_week=report.NewSection(count=len(new_week), rows=list(new_week)),
        ),
        distribution=buckets.distribution(all_rows),
        all_rows=all_rows,
    )


def _scholars_payload():
    """Scholars only: 3 accepted Plumbing, 1 on-hold HVAC, 1 review Plumbing
    (live) plus 2 Plumbing closed this week. Every other community empty."""
    live = [
        _row(number="WO-1", community="Scholars", service_type="Plumbing"),
        _row(number="WO-2", community="Scholars", service_type="Plumbing"),
        _row(
            number="WO-3",
            community="Scholars",
            service_type="Plumbing",
            notes="Tenant home after 5.\nCall first.",
        ),
        _row(
            number="WO-4",
            community="Scholars",
            service_type="HVAC",
            status=wo.STATUS_ON_HOLD,
        ),
        _row(
            number="WO-5",
            community="Scholars",
            service_type="Plumbing",
            status=wo.STATUS_REVIEW,
        ),
    ]
    closed = [
        _row(
            number="WO-6",
            community="Scholars",
            service_type="Plumbing",
            status=wo.STATUS_COMPLETED,
            archived=True,
            labor_total="100.00",
            materials_total="25.00",
        ),
        _row(
            number="WO-7",
            community="Scholars",
            service_type="Plumbing",
            status=wo.STATUS_REVIEW,
            archived=True,
        ),
    ]
    return _payload(
        live=live,
        closed_today=closed[:1],
        closed_week=closed,
        closing=[live[4]],
        by_status={wo.STATUS_REVIEW: 1},
        new_today=[live[0]],
        new_week=live[:2],
    )


def _workbook(payload):
    return openpyxl.load_workbook(io.BytesIO(xlsx.report_xlsx(payload)))


def _cells(sheet):
    """Every cell value as a grid, `None` intact."""
    return [[cell.value for cell in row] for row in sheet.iter_rows()]


def _column(sheet, column, first, last):
    return [sheet.cell(row=row, column=column).value for row in range(first, last + 1)]


def _chart_parts(payload):
    with zipfile.ZipFile(io.BytesIO(xlsx.report_xlsx(payload))) as archive:
        return {
            name: archive.read(name).decode()
            for name in archive.namelist()
            if name.startswith("xl/charts/chart") and name.endswith(".xml")
        }


# --- workbook shape (E6, E7, E10) ------------------------------------------


def test_sheets_are_the_eight_visible_plus_hidden_chart_data_with_data_last():
    workbook = _workbook(_payload())

    assert workbook.sheetnames == [
        "Report",
        "Scholars",
        "Centennial",
        "Commons",
        "Young Hall",
        "Academics",
        "Work Orders",
        "Chart Data",
        "Data",
    ]
    assert workbook.sheetnames == list(xlsx.SHEET_NAMES)
    assert workbook["Chart Data"].sheet_state == "hidden"
    assert all(
        workbook[name].sheet_state == "visible"
        for name in workbook.sheetnames
        if name != "Chart Data"
    )
    assert workbook.active.title == "Report"


def test_tab_colors_and_gridlines_follow_the_house_style():
    workbook = _workbook(_payload())

    assert workbook["Report"].sheet_properties.tabColor.rgb.endswith("C8102E")
    assert workbook["Commons"].sheet_properties.tabColor.rgb.endswith("5A5C60")
    assert workbook["Work Orders"].sheet_properties.tabColor.rgb.endswith("1C1D20")
    assert workbook["Data"].sheet_properties.tabColor.rgb.endswith("B7B9BC")
    assert workbook["Report"].sheet_view.showGridLines is False
    assert workbook["Data"].sheet_view.showGridLines is not False


def test_filename_is_the_covered_day():
    assert xlsx.report_xlsx_filename(_payload()) == "wo-report_2026-08-25.xlsx"


# --- Data (X5, E10) ---------------------------------------------------------


def test_data_sheet_matches_report_csv():
    """The load-bearing pin (X5): the Data sheet is `report_csv`, cell for
    cell, so save-as-CSV from Excel still round-trips through the importer."""
    closed = _row(
        number="WO-1",
        community="North",
        service_type="Plumbing",
        status=wo.STATUS_COMPLETED,
        archived=True,
    )
    closing = _row(number="WO-3", status=wo.STATUS_REVIEW)
    new = _row(number="WO-4")
    payload = _payload(
        live=[closing, new],
        closed_today=[closed],
        closed_week=[
            closed,
            _row(number="WO-2", service_type="HVAC", status=wo.STATUS_COMPLETED, archived=True),
        ],
        closing=[closing],
        new_today=[new],
        new_week=[new],
        by_status={wo.STATUS_REVIEW: 1},
    )

    expected = list(csv.reader(io.StringIO(report.report_csv(payload))))
    actual = [
        ["" if value is None else str(value) for value in row]
        for row in _cells(_workbook(payload)["Data"])
    ]

    assert actual == expected


def test_data_sheet_header_is_the_section_prefixed_export_headers():
    sheet = _workbook(_payload())["Data"]

    assert tuple(cell.value for cell in sheet[1]) == (
        report.CSV_SECTION_HEADER,
    ) + wo.EXPORT_HEADERS
    assert sheet.freeze_panes == "A2"


# --- Work Orders (E4, E5, §4.3) ---------------------------------------------


def test_work_orders_header_has_notes_in_column_c_and_freezes_at_d6():
    sheet = _workbook(_scholars_payload())["Work Orders"]

    assert [cell.value for cell in sheet[5]][:21] == [
        "WORK ORDER",
        "LOCATION",
        "NOTES",
        "BUCKET",
        "STATUS",
        "COMMUNITIES",
        "SERVICE TYPE",
        "PRIORITY",
        "BUILDING",
        "UNIT",
        "SUPERVISOR",
        "TECHNICIANS",
        "MATERIAL LINES",
        "MATERIALS TOTAL",
        "LABOR MINUTES",
        "LABOR TOTAL",
        "TOTAL",
        "CREATED AT",
        "COMPLETED AT",
        "CLOSED AT",
        "SECTIONS",
    ]
    assert sheet["C5"].value == "NOTES"
    assert sheet.column_dimensions["C"].width == 60
    assert sheet.freeze_panes == "D6"
    assert sheet.print_title_rows == "$5:$5"


def test_work_orders_is_the_deduped_population_in_reading_order():
    sheet = _workbook(_scholars_payload())["Work Orders"]

    numbers = _column(sheet, 1, 6, 12)
    assert numbers == ["WO-6", "WO-7", "WO-5", "WO-4", "WO-1", "WO-2", "WO-3"]
    assert len(numbers) == len(set(numbers)) == 7
    assert sheet.cell(row=13, column=1).value is None
    assert [t.ref for t in sheet.tables.values()] == ["A5:U12"]


def test_work_orders_columns_carry_bucket_status_sections_and_real_numbers():
    sheet = _workbook(_scholars_payload())["Work Orders"]
    closed = [cell.value for cell in sheet[6]]
    accepted = [cell.value for cell in sheet[10]]
    ready = [cell.value for cell in sheet[8]]

    assert closed[0] == "WO-6"
    assert closed[3] == "Closed" and closed[4] == "Completed"
    assert closed[12] == 2
    assert closed[13] == Decimal("25.00") and sheet["N6"].number_format == "$#,##0.00"
    assert closed[15] == Decimal("100.00") and closed[16] == Decimal("125.00")
    assert closed[17] == datetime(2026, 8, 25, 14, 0) and sheet["R6"].number_format == "yyyy-mm-dd hh:mm"
    assert closed[19] == datetime(2026, 8, 25, 15, 30)
    assert closed[20] == "closed_today; closed_week"
    assert accepted[3] == "Accepted" and accepted[4] == "Created"
    assert accepted[19] is None
    assert accepted[20] == "new_today; new_week"
    assert ready[3] == "Ready to close" and ready[4] == "Review" and ready[20] == "closing"


def test_work_orders_communities_column_lists_every_membership():
    payload = _payload(
        live=[
            _row(number="WO-1", location="Scholars 3 / Commons annex"),
            _row(number="WO-2", community="Young Hall"),
            _row(number="WO-3"),
        ]
    )
    sheet = _workbook(payload)["Work Orders"]

    assert sheet["F5"].value == "COMMUNITIES"
    assert sheet.column_dimensions["F"].width == 22
    # Memberships, not the raw column (E14): two names for a two-community
    # location, Academics for a location naming none -- never blank.
    assert _column(sheet, 6, 6, 8) == ["Scholars; Commons", "Young Hall", "Academics"]


def test_work_orders_notes_wrap_top_aligned_with_a_capped_row_height():
    sheet = _workbook(_scholars_payload())["Work Orders"]

    note_cell = sheet["C12"]
    assert note_cell.value == "Tenant home after 5.\nCall first."
    assert note_cell.alignment.wrap_text is True
    assert note_cell.alignment.vertical == "top"
    assert sheet.row_dimensions[12].height == 27.0
    assert sheet.row_dimensions[6].height is None


def test_empty_population_renders_a_header_and_an_empty_state_without_a_table():
    sheet = _workbook(_payload())["Work Orders"]

    assert sheet["C5"].value == "NOTES"
    assert sheet["A6"].value == "No live or recently closed work orders."
    assert sheet.tables == {}


# --- Report (§4.1, E11) -----------------------------------------------------


def test_report_title_block_names_the_day_week_and_population():
    sheet = _workbook(_scholars_payload())["Report"]

    assert sheet["A1"].value == "Weekly Work Order Report"
    assert sheet["A2"].value == "Tue, Aug 25, 2026 · week of 2026-08-24 – 2026-08-30 (week to date)"
    assert sheet["A3"].value == "Generated 2026-08-25 17:30 Central"
    assert sheet["A4"].value.startswith("Live work orders as of 2026-08-25 17:30 Central, plus work orders closed 2026-08-24 – 2026-08-30.")
    assert "counted in both" in sheet["A4"].value
    assert sheet.freeze_panes == "A5"


def test_report_kpi_strip_is_open_then_the_three_live_buckets_then_closed():
    sheet = _workbook(_scholars_payload())["Report"]

    assert [cell.value for cell in sheet[xlsx.KPI_ROW]][:5] == [
        "Open work orders",
        "Accepted",
        "In progress",
        "Ready to close",
        "Closed this week",
    ]
    assert [cell.value for cell in sheet[xlsx.KPI_ROW + 1]][:5] == [5, 3, 1, 1, 2]
    assert sheet.cell(row=xlsx.KPI_ROW + 1, column=1).border.bottom.color.rgb.endswith("C8102E")


def test_report_company_status_block_has_counts_and_real_shares():
    sheet = _workbook(_scholars_payload())["Report"]

    assert sheet.cell(row=xlsx.STATUS_ROW, column=1).value == "Company status"
    assert [cell.value for cell in sheet[xlsx.STATUS_ROW + 1]][:3] == ["Bucket", "Count", "Share"]
    rows = [
        [sheet.cell(row=row, column=col).value for col in (1, 2, 3)]
        for row in range(xlsx.STATUS_ROW + 2, xlsx.STATUS_ROW + 6)
    ]
    assert [row[:2] for row in rows] == [
        ["Accepted", 3],
        ["In progress", 1],
        ["Ready to close", 1],
        ["Closed", 2],
    ]
    assert [row[2] for row in rows] == pytest.approx([3 / 7, 1 / 7, 1 / 7, 2 / 7])
    assert sheet.cell(row=xlsx.STATUS_ROW + 2, column=3).number_format == "0.0%"


def test_report_activity_block_matches_the_payload_counts():
    sheet = _workbook(_scholars_payload())["Report"]

    assert sheet.cell(row=xlsx.ACTIVITY_ROW, column=1).value == "Activity"
    # The corner cell is written as None: openpyxl reloads "" as None anyway.
    assert [cell.value for cell in sheet[xlsx.ACTIVITY_ROW + 1]][:3] == [None, "Today", "Week to date"]
    assert [
        [sheet.cell(row=row, column=col).value for col in (1, 2, 3)]
        for row in (xlsx.ACTIVITY_ROW + 2, xlsx.ACTIVITY_ROW + 3)
    ] == [["Closed", 1, 2], ["New", 1, 2]]
    assert sheet.cell(row=xlsx.ACTIVITY_ROW + 4, column=1).value is None


def test_report_activity_block_notes_auto_closed_work_orders():
    payload = _payload(
        closed_today=[_row(status=wo.STATUS_COMPLETED, archived=True)],
        auto_closed_today=1,
    )
    sheet = _workbook(payload)["Report"]

    assert sheet.cell(row=xlsx.ACTIVITY_ROW + 4, column=1).value == (
        "Closed today includes (1 in NetFacilities); this week (0 in NetFacilities)"
    )


def test_report_dollars_block_is_by_primary_community_and_counts_each_row_once():
    week = [
        _row(number="WO-1", community="Centennial", status=wo.STATUS_COMPLETED, archived=True,
             labor_total="100.00", materials_total="10.00"),
        # Named in two communities: attributed to Scholars (first in filter
        # order) and counted once (E14).
        _row(number="WO-2", location="Commons annex / Scholars 3", status=wo.STATUS_COMPLETED,
             archived=True, labor_total="5.00", materials_total="0.00"),
        _row(number="WO-3", community="Centennial", status=wo.STATUS_COMPLETED, archived=True,
             labor_total="50.00", materials_total="20.00"),
        # Nothing named: Academics, never "(no community)".
        _row(number="WO-4", status=wo.STATUS_COMPLETED, archived=True,
             labor_total="1.00", materials_total="0.00"),
    ]

    assert xlsx._community_money(week) == [
        ("Centennial", Decimal("150.00"), Decimal("30.00")),
        ("Scholars", Decimal("5.00"), Decimal("0.00")),
        ("Academics", Decimal("1.00"), Decimal("0.00")),
    ]

    sheet = _workbook(_payload(closed_week=week))["Report"]
    assert sheet.cell(row=xlsx.DOLLARS_ROW, column=1).value == "Dollars closed this week"
    assert [cell.value for cell in sheet[xlsx.DOLLARS_ROW + 1]][:4] == ["Community", "Labor", "Materials", "Total"]
    assert [cell.value for cell in sheet[xlsx.DOLLARS_ROW + 2]][:4] == [
        "Centennial", Decimal("150.00"), Decimal("30.00"), Decimal("180.00"),
    ]
    assert sheet.cell(row=xlsx.DOLLARS_ROW + 2, column=2).number_format == "$#,##0.00"
    # Three money rows (44-46), footnote on the next row (E14).
    assert sheet.cell(row=xlsx.DOLLARS_ROW + 5, column=1).value == xlsx.DOLLARS_FOOTNOTE
    assert xlsx.DOLLARS_FOOTNOTE == "Dollars count a work order under its first community."


def test_report_by_community_table_and_the_ready_to_close_footnote():
    sheet = _workbook(_scholars_payload())["Report"]
    grid = _cells(sheet)
    top = next(index for index, row in enumerate(grid, start=1) if row and row[0] == "By community")

    assert top == xlsx.BY_COMMUNITY_ROW  # fixed (P8): the dollars block is bounded by E14
    assert grid[top][:6] == ["Community", "Accepted", "In progress", "Ready to close", "Closed", "Total"]
    assert grid[top + 1][:6] == ["Scholars", 3, 1, 1, 2, 7]
    assert grid[top + 5][:6] == ["Academics", 0, 0, 0, 0, 0]
    assert [t.displayName for t in sheet.tables.values()] == ["ByCommunity"]
    footnotes = [row[0] for row in grid[top + 6 : top + 9] if row and row[0]]
    assert footnotes[0] == "Ready to close = 0 ready to complete, 0 completed, 1 review."
    assert "counted in both" in footnotes[1]


def test_report_with_nothing_to_plot_shows_the_empty_state_instead_of_a_zero_pie():
    sheet = _workbook(_payload())["Report"]

    assert sheet.cell(row=xlsx.STATUS_ROW, column=7).value == "No live or recently closed work orders."
    assert [cell.value for cell in sheet[xlsx.KPI_ROW + 1]][:5] == [0, 0, 0, 0, 0]
    assert [cell.value for cell in sheet[xlsx.DOLLARS_ROW + 2]][:4] == [
        "(none)", Decimal("0.00"), Decimal("0.00"), Decimal("0.00"),
    ]
