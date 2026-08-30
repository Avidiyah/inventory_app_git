"""The workbook's house style helpers (redesign spec §4.0, E12).

openpyxl cannot read back charts it wrote, so chart assertions go through
`zipfile` over the saved bytes -- the same witness the renderer tests use.
"""

import io
import os
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl import Workbook

from app.services import _xlsx_theme as theme


def _saved(workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _part(data, name):
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return archive.read(name).decode()


def _pie_workbook(**kwargs):
    workbook = Workbook()
    sheet = workbook.active
    source = workbook.create_sheet("Chart Data")
    source.append(["Bucket", "Count"])
    for label, count in (("Accepted", 3), ("In progress", 2), ("Ready to close", 1), ("Closed", 4)):
        source.append([label, count])
    chart = theme.pie_of(
        source,
        title="Status",
        header_row=1,
        last_row=5,
        colors=["9CA3AF", "D97706", "6D28D9", "15803D"],
        **kwargs,
    )
    theme.place(sheet, chart, "B24:D36")
    return workbook


def test_pie_is_fixed_order_fixed_color_and_plots_hidden_cells():
    xml = _part(_saved(_pie_workbook()), "xl/charts/chart1.xml")

    assert 'plotVisOnly val="0"' in xml
    assert 'firstSliceAng val="0"' in xml
    assert xml.count("<dPt>") == 4
    assert xml.index("9CA3AF") < xml.index("D97706") < xml.index("6D28D9") < xml.index("15803D")
    assert 'showPercent val="1"' in xml
    assert 'legendPos val="r"' in xml


def test_pie_can_drop_labels_and_move_the_legend():
    xml = _part(_saved(_pie_workbook(legend="b", percent_labels=False)), "xl/charts/chart1.xml")

    assert 'showPercent val="1"' not in xml
    assert 'legendPos val="b"' in xml


def test_place_fills_exactly_the_named_cells():
    xml = _part(_saved(_pie_workbook()), "xl/drawings/drawing1.xml")

    assert "<twoCellAnchor>" in xml
    assert "<from><col>1</col><colOff>0</colOff><row>23</row>" in xml
    assert "<to><col>4</col><colOff>0</colOff><row>36</row>" in xml


def test_column_chart_uses_brand_then_neutral_and_can_stack():
    workbook = Workbook()
    sheet = workbook.active
    source = workbook.create_sheet("Chart Data")
    source.append(["Community", "Labor", "Materials"])
    source.append(["North", 100, 20])
    chart = theme.column_chart_of(
        source, title="Dollars", header_row=1, last_row=2, stacked=True, y_title="Dollars"
    )
    theme.place(sheet, chart, "G42:M56")

    xml = _part(_saved(workbook), "xl/charts/chart1.xml")

    assert 'grouping val="stacked"' in xml
    assert 'overlap val="100"' in xml
    assert xml.index(theme.BRAND_RED) < xml.index(theme.MUTED)
    assert 'plotVisOnly val="0"' in xml


def test_table_of_writes_a_banded_table_only_when_there_are_rows():
    workbook = Workbook()
    with_rows = workbook.active
    last = theme.table_of(
        with_rows, name="WorkOrders", row=5, headers=("A", "B"), rows=[[1, 2], [3, 4]],
        formats={1: theme.COUNT},
    )
    without = workbook.create_sheet("Empty")
    empty_last = theme.table_of(without, name="Nothing", row=5, headers=("A", "B"), rows=[])

    assert last == 7
    assert empty_last == 5  # the header row is the last row written
    assert [t.displayName for t in with_rows.tables.values()] == ["WorkOrders"]
    assert list(with_rows.tables.values())[0].ref == "A5:B7"
    assert with_rows["B6"].number_format == theme.COUNT
    assert with_rows["A5"].font.bold and with_rows["A5"].font.color.rgb.endswith(theme.WHITE)
    assert without.tables == {}
    assert without["A5"].value == "A"


def test_setup_sheet_applies_the_shared_furniture():
    workbook = Workbook()
    sheet = workbook.active
    theme.setup_sheet(sheet, tab_color=theme.BRAND_RED, freeze="D6", print_title_rows="5:5")

    reloaded = openpyxl.load_workbook(io.BytesIO(_saved(workbook))).active

    assert reloaded.sheet_properties.tabColor.rgb.endswith(theme.BRAND_RED)
    assert reloaded.sheet_view.showGridLines is False
    assert reloaded.freeze_panes == "D6"
    assert reloaded.page_setup.orientation == "landscape"
    assert reloaded.page_setup.fitToWidth == 1
    assert reloaded.sheet_properties.pageSetUpPr.fitToPage is True
    assert reloaded.page_margins.left == 0.5
    assert reloaded.print_title_rows == "$5:$5"  # openpyxl absolutises on reload


def test_setup_sheet_can_leave_gridlines_on_for_the_data_sheet():
    workbook = Workbook()
    theme.setup_sheet(workbook.active, tab_color=theme.TAB_GRAY, freeze="A2", gridlines=True)

    reloaded = openpyxl.load_workbook(io.BytesIO(_saved(workbook))).active

    assert reloaded.sheet_view.showGridLines is not False


def test_title_block_section_and_kpi_styles():
    workbook = Workbook()
    sheet = workbook.active
    theme.title_block(sheet, "Scholars", ["Weekly status", "caveat"])
    theme.section(sheet, 5, "Status", span=3)
    theme.kpi(sheet, 6, 2, "Accepted", 12)

    assert sheet["A1"].font.size == 18 and sheet["A1"].font.color.rgb.endswith(theme.BRAND_RED)
    assert sheet["A2"].font.italic and sheet["A2"].font.size == 9
    assert sheet["A5"].font.bold and sheet["C5"].border.bottom.style == "thin"
    assert sheet["D5"].border.bottom.style is None
    assert sheet["B6"].value == "Accepted" and sheet["B7"].value == 12
    assert sheet["B7"].border.bottom.color.rgb.endswith(theme.BRAND_RED)
    assert sheet["B7"].number_format == theme.COUNT


def test_notes_row_height_caps_at_four_lines():
    assert theme.notes_row_height(None) is None
    assert theme.notes_row_height("short") is None
    assert theme.notes_row_height("one\ntwo") == 27.0
    assert theme.notes_row_height("x" * 200) == 54.0
    assert theme.notes_row_height("\n".join(["line"] * 10)) == 54.0
